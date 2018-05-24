import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import pro_qcn_xmlfile
import time

                
def excel(file):                   
    #加载nv_updata_list文件夹内的nv更新表                   
    wb = load_workbook(file)
    sheet = wb[wb.sheetnames[0]]
    n = 2
    Acellvalue = ''
    valuedict = {}
    Fcelldict = {}
    Gcelldict = {}
    
    #获取xlsx文件内的数据及相关单元格位置
    while 1 :
        Acellvalue = sheet['A'+str(n)].value
        Ecellvalue = sheet['E'+str(n)].value
        if Acellvalue == None:
            break
        valuedict[Acellvalue] = Ecellvalue
        Fcelldict[Acellvalue] = 'F'+str(n)   #显示结果的单元格位置
        Gcelldict[Acellvalue] = 'G'+str(n)   #显示检查时间的单元格位置
        n = n + 1
    
    nvdic = pro_qcn_xmlfile.qcn_xml()

                
                
    redFill = PatternFill(start_color='FFFF0000',
                       end_color='FFFF0000',
                       fill_type='solid')
    greenFill = PatternFill(start_color='FF00FF00',
                       end_color='FF00FF00',
                       fill_type='solid')   
            
    for key in valuedict:
        try:
            if valuedict[key] == nvdic[key]:
                print(key, 'OK')
                sheet[Fcelldict[key]] = 'OK'
                sheet[Fcelldict[key]].fill  = greenFill
                sheet[Gcelldict[key]] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) 
            else:
                print(key, 'NOK')
                sheet[Fcelldict[key]] = 'NOK'
                sheet[Fcelldict[key]].fill  = redFill
                sheet[Gcelldict[key]] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
        except KeyError:
                print(key, 'NOK')
                sheet[Fcelldict[key]] = 'NOK'
                sheet[Fcelldict[key]].fill  = redFill
                sheet[Gcelldict[key]] = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()) 
    OK = 1
    while OK:
        try:       
            wb.save(file) 
            OK = 0     
        except PermissionError as e:
            print (e)
            input('check whether the excel file is opened, please close it and try')

if __name__ == '__main__'    :
    path = "./nv_updata_list"
    files= os.listdir(path)
    file = files[0]
    filetype = file.split('.')[-1]
    if filetype == 'xlsx' or filetype == 'xls':
        file = "./nv_updata_list/"+file
        excel(file)
    else:
        print('no excel file')        
